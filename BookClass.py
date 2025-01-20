import openpyxl
from BoyClass import boy

class book():
    #Я думаю, что объекты для мальчиков со временем складываются в кучу и в конечном итоге разобьются.
    def __init__(self):
        self.Sheetlength = 0
        self.Boylist = []
    def useExternalBoylist(self,boylist):
        self.Boylist = boylist
    def gen(self,adress,superBoyList):
        self.data = openpyxl.load_workbook(adress).active
        self.Sheetlength = 0
        self.Boylist = []

        for cols in self.data.iter_cols(5,self.data.max_column):
            if str(cols[0].value)[0] == "P":
                break
            else:
                self.Sheetlength += 1 
        for rowindex, row in enumerate(self.data.iter_rows(1, self.data.max_row)):
            if row[1].value == 'Key' or row[1].value == 'KEY':
                break
            if row[1].value != None:
                fiveScores = []
                fiveDates = []
                fiveTypes = []
                boyname = str(row[1].value) + " " + str(row[2].value)
                for col in self.data.iter_cols(5,self.Sheetlength):
                    value = col[rowindex].value
                    if value != None:
                        value = str(value).strip()
                        if value != '':
                            fiveScores.append(int(value))
                            fiveDates.append(col[rowindex+3].value) 
                            fiveTypes.append(col[rowindex+2].value)
        # делает классы на 5

                tenScores = []
                tenDates = []
                tenTypes = []

                for col in self.data.iter_cols(5,self.Sheetlength):
                    value = col[rowindex + 1].value
                    if value != None:
                        value = str(value).strip()
                        if value != '':
                            tenScores.append(int(value))
                            tenDates.append(col[rowindex+3].value)
                            tenTypes.append(col[rowindex+2].value)
                
        # делает классы на 10
                xtras = []

                for col in self.data.iter_cols(self.Sheetlength+9+4,4+18+self.Sheetlength):
                    xtras.append(col[rowindex].value)
                    
            
                fivebull = [fiveScores,fiveDates,fiveTypes]
                tenbull = [tenScores,tenDates,tenTypes]

    
                self.Boylist.append(boy(boyname,fivebull,tenbull,xtras))
        superBoyList.extend(self.Boylist)
        return superBoyList
    def findBoy(self,name):
        for index, boy in enumerate(self.Boylist):
            if str(boy.name).lower() == name.lower():
                return index
        return -1
    
    def getFiveAvg(self,index):
        return self.Boylist[index].getFiveAvg()
    def getTenAvg(self,index):
        return self.Boylist[index].getTenAvg()
    def plotFiveAvg(self,index):
        self.Boylist[index].plot(self.Boylist[index].fiveBull[0],self.Boylist[index].fiveBull[1])
    def plotTenAvg(self,index):
        self.Boylist[index].plot(self.Boylist[index].tenBull[0],self.Boylist[index].tenBull[1])
    def getTargetLeaderboard(self,type):
        avgFiveList = []
        nameList = []
        for boyindex, boy in enumerate(self.Boylist):
            if type == 5:
                avgFiveList.append(self.getFiveAvg(boyindex))
            elif type == 10:
                avgFiveList.append(self.getTenAvg(boyindex))
            nameList.append(boy.name)

        paired = list(zip(nameList, avgFiveList))
        sortedPaired = sorted(paired, key=lambda x: x[1],reverse=True)

        names = [name for name, i in sortedPaired]
        numbers = [num for i, num in sortedPaired]

        return names,numbers
    def returnBoyList(self):
        return self.Boylist
    def displayMisc(self,index):
        self.Boylist[index].displayMisc()

    
        